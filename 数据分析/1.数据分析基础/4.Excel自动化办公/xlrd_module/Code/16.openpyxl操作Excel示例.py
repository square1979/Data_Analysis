from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

# 抓取工作表
ws = wb.active
# 数据分配给单元格
ws['A1'] = '行文脉络'
# 列添加元素
ws.append([1, 2, 3])
ws['A3'] = datetime.now()
wb.save('test.xlsx')
