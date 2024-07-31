from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook

# 数字变为字母
print(get_column_letter(20))
# 字母转为数字
print(column_index_from_string('p'))

# 删除工作表
wb = load_workbook('lucky.xlsx')
# 使用remove方法需要先获取到工作表的名称
# sheet = wb.get_sheet_by_name('num1')
# wb.remove(sheet)
# 使用del删除工作表
del wb['num']
wb.save('lucky.xlsx')
