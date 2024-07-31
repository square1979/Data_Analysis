import datetime
from time import time

from openpyxl import Workbook

# 新建文件
wb = Workbook()
# 新建工作表
ws = wb.create_sheet('演示')
# 第一行输入数据
ws.append(['time', 'title'])
# 输入内容【500行数据】
for i in range(500):
    TIME = datetime.datetime.now().strftime('%H:%M:%S')
    TITLE = str(time())
    ws.append([TIME, TITLE])

# 获取最大行
print(ws.max_row)
print(ws.max_column)
# 将上述内容打印在控制台上
for col in ws.columns:
    for cell in col:
        print(cell, end="")
    print()
wb.save('演示.xlsx')
