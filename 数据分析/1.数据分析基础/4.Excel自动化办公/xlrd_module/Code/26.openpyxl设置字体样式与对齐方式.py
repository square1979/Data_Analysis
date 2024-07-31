from openpyxl import Workbook
from openpyxl.styles import colors, Font, Alignment

wb = Workbook()
ws = wb.active
# 创建字体对象
font = Font(size=24, italic=True, bold=True, color=colors.COLOR_INDEX[19])
# 给单元格A1设置内容
ws['A1'] = 'lucky'
# 给A1设置样式
ws['A1'].font = font
# 设置字体对齐方式
ws['A2'] = '玛尔扎哈'
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
wb.save('新建.xlsx')
