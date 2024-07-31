from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 设置第二行的行高
ws.row_dimensions[2].height = 40
# 设置c列的列宽
ws.column_dimensions['C'].width = 50
# 合并单元格
sheet = wb.create_sheet('新建')
sheet.merge_cells('A1:D1')
sheet.merge_cells('A5:D6')
sheet.unmerge_cells('A5:D6')
wb.save('新建.xlsx')
