from openpyxl import load_workbook
wb = load_workbook('test.xlsx')
# 获取工作表
ws1 = wb.active
ws1['A1'] = 'lucky'
# 创建工作表并写入数据
# ws2 = wb.create_sheet('sheet_2')
# ws2['A1'] = 'lucky_sheet_2'
# 通过sheet名称，获取sheet
ws3 = wb['sheet_2']
ws3['A2'] = 'lucky_sheet_3'
# 老方法根据名字写入单元格数据
ws4 = wb.get_sheet_by_name('sheet_21')
ws4['A5'] = '星星之火可以燎原'
# 查看表格名称
sheets = wb.sheetnames
for sheet in sheets:
    print(sheet)

wb.save('test.xlsx')
