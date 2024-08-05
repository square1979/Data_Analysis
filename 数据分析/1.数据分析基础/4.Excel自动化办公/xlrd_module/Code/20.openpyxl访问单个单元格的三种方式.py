from openpyxl import Workbook

# 实例化workbook
wb = Workbook()
# 激活工作表
ws = wb.active
# 方法1 通过单元格名称访问
ws['A10'] = "星星之火可以燎原"
# 方法2 通过行列定位单元格访问
ws.cell(row=2, column=2, value="星星之火可以燎原！")
# 方法3 使用循环写入数据
for row in range(1, 10):
    for col in range(1, row+1):
        ws.cell(row=row, column=col, value=f'{row}*{col}={str(row * col)}')


wb.save('星星之火.xlsx')
