import openpyxl
from openpyxl import Workbook

wb = Workbook()
wb1 = openpyxl.load_workbook('lucky.xlsx')

ws = wb.active
# 开始设置颜色
ws.sheet_properties.tabColor = '99FFFF'
wb.save('新建.xlsx')
# 获取最大行/最大列

ws1 = wb1.active
print(ws1.max_row)
print(ws1.max_column)
