from openpyxl import load_workbook
from datetime import datetime

# 加载现有数据
wb = load_workbook('lucky.xlsx')
# 在工作簿最后创建工作表，默认最后
# wb = wb.create_sheet('my sheet')
# 在最前面创建工作表
wb.create_sheet('num1', 2)
ws = wb.active

# 向A1单元格写入内容
ws['A1'] = 'lucky'
# 向下一行写入数据
ws.append(['l', 'lu', 'luc', 'luck', 'lucky'])
# A3单元格写入时间，格式自动转换
ws['A3'] = datetime.now().strftime('%Y-%m-%d')

# 创建工作表


# 保存数据
wb.save('lucky.xlsx')
