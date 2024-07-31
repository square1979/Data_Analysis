from openpyxl import Workbook
from openpyxl import load_workbook

# 创建实例化对象
wb1 = Workbook()
ws1 = wb1.active

# 打开现有文件
wb2 = load_workbook('test.xlsx')
