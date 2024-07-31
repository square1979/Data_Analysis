from openpyxl import Workbook
# 新建工作簿
wb = Workbook()
# 激活工作表
ws = wb.active
# 通过切片访问多个单元格
c_range0 = ws['A1:C2']
print(c_range0)
print('--------------')
# 通过行列访问多个单元格
c_range1 = ws['C']
print(c_range1)
print('--------------')
c_range2 = ws['C:D']
print(c_range2)
print('--------------')
c_range3 = ws[5]
print(c_range3)
print('--------------')
c_range4 = ws[5:10]
print(c_range4)
print('--------------')
