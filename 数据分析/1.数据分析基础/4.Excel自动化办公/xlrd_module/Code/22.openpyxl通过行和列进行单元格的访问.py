from openpyxl import load_workbook

wb = load_workbook('lucky.xlsx')
ws = wb.active
# print(list(ws.iter_rows(min_row=1, max_row=2, max_col=5)))
i = 1
# 通过指定范围(⾏ → ⾏)
"""
for row in ws.iter_rows(min_row=1, max_row=7, max_col=5):
    # print(row)
    for col in row:
        # print(type(col.value), col)
        col.value = col.value + str(i)
    i += 1
"""
# 通过指定范围（列 → 列）
for col in ws.iter_cols(min_row=1, max_col=5, max_row=7):
    for row in col:
        # print(row)
        row.value = row.value + str(i)
    i += 1

wb.save('lucky.xlsx')
